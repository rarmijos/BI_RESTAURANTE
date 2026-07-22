# =====================================================
# generar_informe_restaurante.py
# DESCRIPCIÓN: Genera Informe_Restaurante.xlsx
# EXACTAMENTE IGUAL a las tablas de Adminer
# =====================================================

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =====================================================
# CONFIGURACIÓN
# =====================================================
NUM_COMPRAS = 500
FECHA_INICIO = datetime(2025, 1, 1)
FECHA_FIN = datetime(2026, 12, 31)
PROB_INCONSISTENCIA = 0.20

# =====================================================
# 1. DATOS DE PROVEEDORES (8 registros)
# =====================================================
proveedores_data = [
    ("D001", "Distribuidora Andina Farma", "Quito", "Carolina Salazar", "0998001001", "carolina@andinafarma.ec", "Crédito 30 días"),
    ("D002", "Medisupply Ecuador", "Guayaquil", "Pablo Andrade", "0998001002", "pablo@medisupply.ec", "Crédito 45 días"),
    ("D003", "FarmaLogística Nacional", "Quito", "Andrea Pozo", "0998001003", "andrea@farmalogistica.ec", "Contado"),
    ("D004", "Distribuciones VitalMed", "Cuenca", "Diego Zambrano", "0998001004", "diego@vitalmed.ec", "Crédito 30 días"),
    ("D005", "Comercial PharmaSur", "Quito", "Lucía Montalvo", "0998001005", "lucia@pharmasur.ec", "Crédito 15 días"),
    ("D006", "Laboratorios y Distribuciones Quito", "Quito", "Fernando Lema", "0998001006", "fernando@ldquito.ec", "Crédito 60 días"),
    ("D007", "Importadora Salud Total", "Manta", "Gabriela Vera", "0998001007", "gabriela@saludtotal.ec", "Crédito 30 días"),
    ("D008", "Medicamentos Express", "Ambato", "Carlos Nieto", "0998001008", "carlos@medexpress.ec", "Contado"),
]

# =====================================================
# 2. DATOS DE INSUMOS (20 registros)
# =====================================================
insumos_data = [
    ("INS001", "Lomo de Res Premium", "Cárnicos", "Premium", "Kg", 4, 5, "México"),
    ("INS002", "Pechuga de Pollo", "Aves", "Filete", "Kg", 2, 4, "México"),
    ("INS003", "Tomate", "Verduras", "Entero", "Kg", 12, 7, "México"),
    ("INS004", "Cebolla", "Verduras", "Entera", "Kg", 15, 15, "México"),
    ("INS005", "Filete de Salmón", "Pescados", "Filete", "Kg", 0, 3, "Chile"),
    ("INS006", "Camarón", "Mariscos", "Entero", "Kg", -1, 3, "Ecuador"),
    ("INS007", "Pollo Entero", "Aves", "Entero", "Kg", 2, 4, "México"),
    ("INS008", "Diezmillo de Res", "Cárnicos", "Estándar", "Kg", 4, 4, "México"),
    ("INS009", "Zanahoria", "Verduras", "Entera", "Kg", 14, 10, "México"),
    ("INS010", "Lechuga", "Verduras", "Entera", "Pieza", 8, 3, "México"),
    ("INS011", "Costilla de Cerdo", "Cárnicos", "Estándar", "Kg", 4, 4, "México"),
    ("INS012", "Pulpa de Cerdo", "Cárnicos", "Premium", "Kg", 4, 5, "México"),
    ("INS013", "Muslo de Pollo", "Aves", "Con hueso", "Kg", 2, 4, "México"),
    ("INS014", "Ala de Pollo", "Aves", "Entera", "Kg", 2, 3, "México"),
    ("INS015", "Lomo de Atún", "Pescados", "Lomo", "Kg", 0, 3, "Ecuador"),
    ("INS016", "Pimiento Morrón", "Verduras", "Entero", "Kg", 10, 5, "México"),
    ("INS017", "Espinaca", "Verduras", "Hoja", "Kg", 6, 3, "México"),
    ("INS018", "Brócoli", "Verduras", "Ramo", "Kg", 8, 4, "México"),
    ("INS019", "Aguacate", "Verduras", "Entero", "Kg", 15, 4, "México"),
    ("INS020", "Champiñón", "Verduras", "Entero", "Kg", 6, 3, "México"),
]

# =====================================================
# 3. DATOS DE SUCURSALES (4 registros - RESTAURANTES)
# =====================================================
sucursales_data = [
    ("REST-001", "Restaurante El Buen Sabor", "Centro", "Quito", "Av. Amazonas 123 y Patria"),
    ("REST-002", "Restaurante La Casona", "Norte", "Quito", "Av. 10 de Agosto N45-20"),
    ("REST-003", "Restaurante Delicias del Mar", "Sur", "Quito", "Av. Maldonado S25-10"),
    ("REST-004", "Restaurante Valle Verde", "Valles", "Sangolquí", "Av. Calderón 101"),
]

# =====================================================
# 4. GENERAR COMPRAS CON INCONSISTENCIAS
# =====================================================
def generar_fecha_aleatoria(inicio, fin):
    dias = (fin - inicio).days
    return inicio + timedelta(days=random.randint(0, dias))

def generar_inconsistencia(tipo_error, compra):
    if tipo_error == 'proveedor_nulo':
        compra['id_proveedor'] = None
    elif tipo_error == 'nombre_vacio':
        compra['nombre_proveedor'] = ''
    elif tipo_error == 'categoria_invalida':
        compra['categoria_insumo'] = 'XX'
    elif tipo_error == 'cantidad_negativa':
        compra['cantidad_comprada'] = -abs(random.randint(1, 50))
    elif tipo_error == 'precio_negativo':
        compra['costo_unitario'] = -abs(round(random.uniform(1, 15), 4))
    elif tipo_error == 'descuento_excesivo':
        compra['descuento_porcentaje'] = random.choice([50, 75, 90, 100])
    elif tipo_error == 'sucursal_invalida':
        compra['id_sucursal'] = 'XXX-999'
    return compra

def generar_compra(i, forzar_inconsistencia=False):
    proveedor = random.choice(proveedores_data)
    insumo = random.choice(insumos_data)
    sucursal = random.choice(sucursales_data)
    fecha_compra = generar_fecha_aleatoria(FECHA_INICIO, FECHA_FIN)
    
    cantidad = round(random.uniform(10, 200), 0)
    costo = round(random.uniform(1.5, 15), 4)
    descuento = random.choice([0, 2.5, 5, 7.5, 10, 15])
    costo_descuento = round(costo * (1 - descuento/100), 4)
    valor_neto = round(cantidad * costo_descuento, 2)
    
    estado = random.choice(['RECIBIDO', 'PENDIENTE', 'PARCIAL'])
    
    compra = {
        'id_compra': f'C-{i+1:04d}',
        'fecha_compra': fecha_compra.strftime('%Y-%m-%d'),
        'id_proveedor': proveedor[0],
        'nombre_proveedor': proveedor[1],
        'id_insumo': insumo[0],
        'nombre_insumo': insumo[1],
        'categoria_insumo': insumo[2],
        'cantidad_comprada': int(cantidad),
        'costo_unitario': float(costo_descuento),
        'descuento_porcentaje': float(descuento),
        'valor_neto': float(valor_neto),
        'id_sucursal': sucursal[0],
        'nombre_sucursal': sucursal[1],
        'estado_recepcion': estado
    }
    
    if forzar_inconsistencia:
        tipo_error = random.choice([
            'proveedor_nulo', 'nombre_vacio', 'categoria_invalida',
            'cantidad_negativa', 'precio_negativo', 'descuento_excesivo',
            'sucursal_invalida'
        ])
        compra = generar_inconsistencia(tipo_error, compra)
    
    return compra

# Generar compras
compras = []
for i in range(NUM_COMPRAS):
    if random.random() < PROB_INCONSISTENCIA:
        compra = generar_compra(i, forzar_inconsistencia=True)
    else:
        compra = generar_compra(i, forzar_inconsistencia=False)
    compras.append(compra)

# =====================================================
# 5. CREAR DATAFRAMES (EXACTAMENTE IGUAL A ADMINER)
# =====================================================

# Hoja 1: excel_compras
df_compras = pd.DataFrame(compras)

# Hoja 2: excel_proveedores
df_proveedores = pd.DataFrame(proveedores_data, 
    columns=['id_proveedor', 'nombre_proveedor', 'ciudad', 'contacto_ventas', 'telefono', 'correo', 'condicion_pago'])

# Hoja 3: excel_insumos
df_insumos = pd.DataFrame(insumos_data,
    columns=['id_insumo', 'nombre_insumo', 'categoria', 'tipo', 'unidad_medida', 'temperatura_requerida', 'vida_util_dias', 'origen'])

# Hoja 4: excel_sucursales
df_sucursales = pd.DataFrame(sucursales_data,
    columns=['id_sucursal', 'nombre_sucursal', 'zona', 'ciudad', 'direccion'])

# Hoja 5: excel_resumen_mensual
df_compras['fecha'] = pd.to_datetime(df_compras['fecha_compra'])
df_compras['anio'] = df_compras['fecha'].dt.year
df_compras['mes'] = df_compras['fecha'].dt.month

resumen_mensual = df_compras.groupby(['anio', 'mes']).agg(
    compras=('id_compra', 'count'),
    unidades=('cantidad_comprada', 'sum'),
    valor_neto=('valor_neto', 'sum')
).reset_index()

resumen_mensual = resumen_mensual.rename(columns={'compras': 'compras', 'unidades': 'unidades', 'valor_neto': 'valor_neto'})

# =====================================================
# 6. FUNCIÓN PARA APLICAR FORMATO
# =====================================================
def aplicar_formato_excel(archivo):
    wb = load_workbook(archivo)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
    
    wb.save(archivo)

# =====================================================
# 7. GUARDAR EXCEL
# =====================================================
NOMBRE_ARCHIVO = 'Informe_Restaurante.xlsx'

with pd.ExcelWriter(NOMBRE_ARCHIVO, engine='openpyxl') as writer:
    df_compras.to_excel(writer, sheet_name='excel_compras', index=False)
    df_proveedores.to_excel(writer, sheet_name='excel_proveedores', index=False)
    df_insumos.to_excel(writer, sheet_name='excel_insumos', index=False)
    df_sucursales.to_excel(writer, sheet_name='excel_sucursales', index=False)
    resumen_mensual.to_excel(writer, sheet_name='excel_resumen_mensual', index=False)

aplicar_formato_excel(NOMBRE_ARCHIVO)

# =====================================================
# 8. ESTADÍSTICAS
# =====================================================
print("="*60)
print("📊 INFORME_RESTAURANTE.XLSX GENERADO")
print("="*60)
print(f"📁 Archivo: {NOMBRE_ARCHIVO}")
print(f"📊 excel_compras: {len(df_compras)} registros")
print(f"📊 excel_proveedores: {len(df_proveedores)} registros")
print(f"📊 excel_insumos: {len(df_insumos)} registros")
print(f"📊 excel_sucursales: {len(df_sucursales)} registros")
print(f"📊 excel_resumen_mensual: {len(resumen_mensual)} registros")

print("\n" + "="*60)
print("🔍 INCONSISTENCIAS GENERADAS")
print("="*60)
print(f"• Proveedor nulo: {df_compras['id_proveedor'].isna().sum()}")
print(f"• Nombre proveedor vacío: {df_compras[df_compras['nombre_proveedor'] == ''].shape[0]}")
print(f"• Categoría inválida (XX): {df_compras[df_compras['categoria_insumo'] == 'XX'].shape[0]}")
print(f"• Cantidad negativa: {df_compras[df_compras['cantidad_comprada'] < 0].shape[0]}")
print(f"• Precio negativo: {df_compras[df_compras['costo_unitario'] < 0].shape[0]}")
print(f"• Descuento excesivo (>50%): {df_compras[df_compras['descuento_porcentaje'] > 50].shape[0]}")
print(f"• Sucursal inválida (XXX-999): {df_compras[df_compras['id_sucursal'] == 'XXX-999'].shape[0]}")

total = df_compras[
    df_compras['id_proveedor'].isna() | 
    (df_compras['nombre_proveedor'] == '') | 
    (df_compras['categoria_insumo'] == 'XX') | 
    (df_compras['cantidad_comprada'] < 0) | 
    (df_compras['costo_unitario'] < 0) | 
    (df_compras['descuento_porcentaje'] > 50) | 
    (df_compras['id_sucursal'] == 'XXX-999')
].shape[0]

print(f"\n📊 Total registros con inconsistencias: {total}")
print(f"📊 Porcentaje: {round(total/NUM_COMPRAS*100, 2)}%")
print("\n✅ ¡Excel generado correctamente!")